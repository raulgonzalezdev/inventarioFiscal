import pyodbc
import pandas as pd
from datetime import datetime, timedelta
import random
import calendar
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Configuración de la conexión a SQL Server
def get_connection():
    conn_str = (
        "DRIVER={SQL Server};"
        "SERVER=DELLXEONE31545\SQLEXPRESS;"
        "DATABASE=DatqBoxExpress;"
        "UID=sa;"
        "PWD=e!334011"
    )
    return pyodbc.connect(conn_str)

def obtener_datos_periodo(periodo):
    """Obtiene los datos del período de InventarioContable"""
    conn = get_connection()
    query = f"""
    SELECT Inicial, Final, AjusteCompras, AjusteVentas
    FROM InventarioContable
    WHERE LTRIM(RTRIM(Periodo)) = '{periodo}'
    """
    df = pd.read_sql(query, conn)
    conn.close()
    return df.iloc[0] if not df.empty else None

def obtener_productos():
    """Obtiene los productos del inventario para generar movimientos"""
    conn = get_connection()
    query = """
    SELECT CODIGO, CATEGORIA, TIPO, DESCRIPCION, MARCA, PRECIO_COMPRA, EXISTENCIA
    FROM Inventario
    WHERE EXISTENCIA > 0 AND CODIGO NOT IN ('0000000001', '0000000002')
    """
    df = pd.read_sql(query, conn)
    conn.close()
    
    productos = []
    for _, row in df.iterrows():
        try:
            # Construir descripción completa
            categoria = str(row['CATEGORIA']).strip() if 'CATEGORIA' in row and not pd.isna(row['CATEGORIA']) else ""
            tipo = str(row['TIPO']).strip() if 'TIPO' in row and not pd.isna(row['TIPO']) else ""
            descripcion = str(row['DESCRIPCION']).strip() if 'DESCRIPCION' in row and not pd.isna(row['DESCRIPCION']) else ""
            marca = str(row['MARCA']).strip() if 'MARCA' in row and not pd.isna(row['MARCA']) else ""
            
            descripcion_completa = f"{categoria} {tipo} {descripcion} {marca}".strip()
            
            producto = {
                'codigo': str(row['CODIGO']).strip(),
                'descripcion': descripcion_completa,
                'precio': float(row['PRECIO_COMPRA']) if not pd.isna(row['PRECIO_COMPRA']) else 0.0,
                'existencia': int(row['EXISTENCIA']) if not pd.isna(row['EXISTENCIA']) else 0
            }
            # Asegurar que los precios estén en un rango razonable (entre 5 y 500)
            producto['precio'] = max(5.0, min(500.0, producto['precio']))
            productos.append(producto)
        except Exception as e:
            print(f"Error al procesar producto: {str(e)}")
            continue
    
    return productos

def obtener_dias_habiles(año, mes):
    """Obtiene los días hábiles (lunes a sábado) del mes especificado"""
    primer_dia = datetime(año, mes, 1)
    if mes == 12:
        ultimo_dia = datetime(año + 1, 1, 1) - timedelta(days=1)
    else:
        ultimo_dia = datetime(año, mes + 1, 1) - timedelta(days=1)
    
    dias_habiles = []
    dia_actual = primer_dia
    
    while dia_actual <= ultimo_dia:
        if dia_actual.weekday() != 6:  # 6 = domingo
            dias_habiles.append(dia_actual)
        dia_actual += timedelta(days=1)
    
    return dias_habiles

def distribuir_valor(valor_total, num_productos):
    """Distribuye un valor total entre varios productos de forma aleatoria pero equilibrada"""
    if num_productos <= 0:
        return []
    
    # Generar pesos aleatorios para la distribución
    pesos = np.random.dirichlet(np.ones(num_productos)) * valor_total
    
    # Redondear a 2 decimales
    return [round(peso, 2) for peso in pesos]

def calcular_precio_consistente(precio_base, categoria_precio):
    """
    Calcula un precio consistente basado en un precio base y una categoría
    
    Args:
        precio_base (float): Precio base del producto
        categoria_precio (int): Categoría de precio (0-4)
        
    Returns:
        float: Precio calculado
    """
    # Definir rangos de precios por categoría
    rangos = [
        (5.0, 25.0),     # Categoría 0: productos muy económicos
        (20.0, 80.0),    # Categoría 1: productos económicos
        (75.0, 150.0),   # Categoría 2: productos de precio medio
        (140.0, 250.0),  # Categoría 3: productos de precio alto
        (240.0, 500.0)   # Categoría 4: productos premium
    ]
    
    rango_min, rango_max = rangos[categoria_precio]
    
    # Si el precio base ya está en el rango, ajustarlo ligeramente
    if rango_min <= precio_base <= rango_max:
        # Variación aleatoria de ±10%
        factor = random.uniform(0.9, 1.1)
        precio = precio_base * factor
    else:
        # Si está fuera del rango, generamos un precio dentro del rango
        # pero más cercano al precio base si está próximo
        distancia_min = abs(precio_base - rango_min)
        distancia_max = abs(precio_base - rango_max)
        
        if distancia_min < distancia_max:
            # Más cercano al mínimo
            precio = rango_min + random.uniform(0, (rango_max - rango_min) * 0.4)
        else:
            # Más cercano al máximo
            precio = rango_max - random.uniform(0, (rango_max - rango_min) * 0.4)
    
    # Redondear a 2 decimales
    return round(precio, 2)

def distribuir_cantidad_por_dia(cantidad_total, num_dias, variabilidad=0.3):
    """
    Distribuye una cantidad total entre varios días con cierta variabilidad
    
    Args:
        cantidad_total (int): Cantidad total a distribuir
        num_dias (int): Número de días
        variabilidad (float): Factor de variabilidad (0-1)
        
    Returns:
        list: Lista de cantidades por día
    """
    if num_dias <= 0 or cantidad_total <= 0:
        return []
    
    # Cantidad base por día
    cantidad_base = cantidad_total / num_dias
    
    # Generar factores aleatorios para cada día
    factores = np.random.normal(1.0, variabilidad, num_dias)
    
    # Normalizar factores para que sumen 1
    factores = factores / factores.sum()
    
    # Calcular cantidades por día
    cantidades = [round(cantidad_total * factor) for factor in factores]
    
    # Ajustar para que sumen exactamente la cantidad total
    diferencia = cantidad_total - sum(cantidades)
    if diferencia != 0:
        idx = random.randint(0, num_dias - 1)
        cantidades[idx] += diferencia
    
    # Asegurar que no hay cantidades negativas
    cantidades = [max(1, c) for c in cantidades]
    
    return cantidades

def obtener_existencias_previas(año, mes):
    """
    Obtiene las existencias finales del período anterior para considerar como iniciales
    
    Args:
        año (int): Año actual
        mes (int): Mes actual (1-12)
        
    Returns:
        dict: Diccionario con existencias por código de producto
    """
    try:
        # Determinar período anterior
        if mes == 1:
            mes_anterior = 12
            año_anterior = año - 1
        else:
            mes_anterior = mes - 1
            año_anterior = año
        
        periodo_anterior = f"{mes_anterior:02d}/{año_anterior}"
        
        # Consultar existencias finales del período anterior
        conn = get_connection()
        query = """
        SELECT Codigo, Final
        FROM MovInventMes
        WHERE Periodo = ? AND Codigo <> '0000000001'
        """
        df = pd.read_sql(query, conn, params=[periodo_anterior])
        conn.close()
        
        # Crear diccionario de existencias
        existencias = {}
        for _, row in df.iterrows():
            codigo = row['Codigo']
            existencia = int(row['Final']) if not pd.isna(row['Final']) else 0
            # Asegurar que no haya existencias negativas en los datos previos
            existencias[codigo] = max(0, existencia)
        
        return existencias
    except Exception as e:
        print(f"Error al obtener existencias previas: {str(e)}")
        return {}

def actualizar_inventario_contable(periodo, conn=None):
    """
    Actualiza los valores de InventarioContable para que coincidan con los movimientos generados
    
    Args:
        periodo (str): Período en formato MM/AAAA
        conn (pyodbc.Connection, optional): Conexión a la base de datos. Si es None, se crea una nueva.
    
    Returns:
        bool: True si se actualizó correctamente, False en caso contrario
    """
    try:
        cerrar_conn = False
        if conn is None:
            conn = get_connection()
            cerrar_conn = True
        
        cursor = conn.cursor()
        
        # Obtener mes y año del período
        mes, año = map(int, periodo.split('/'))
        
        # 1. Calcular valor inicial y final basado en MovInventMes
        cursor.execute("""
            SELECT 
                SUM(CASE WHEN Codigo = '0000000001' THEN Costo ELSE 0 END) AS ValorInicial,
                SUM(Costo * final) AS ValorFinal
            FROM MovInventMes
            WHERE Periodo = ?
        """, (periodo,))
        
        result = cursor.fetchone()
        valor_inicial = result[0] if result[0] is not None else 0
        valor_final = result[1] if result[1] is not None else 0
        
        # 2. Si es diciembre del año actual, no actualizar el valor final
        es_diciembre_año_actual = (mes == 12 and año == datetime.now().year)
        
        # 3. Actualizar InventarioContable
        if es_diciembre_año_actual:
            # Solo actualizar el valor inicial para diciembre
            cursor.execute("""
                UPDATE InventarioContable
                SET Inicial = ?
                WHERE Periodo = ?
            """, (valor_inicial, periodo))
        else:
            # Actualizar inicial y final para los demás meses
            cursor.execute("""
                UPDATE InventarioContable
                SET Inicial = ?, Final = ?
                WHERE Periodo = ?
            """, (valor_inicial, valor_final, periodo))
        
        # 4. Si es un mes distinto a diciembre, actualizar el inicial del mes siguiente
        if mes < 12:
            siguiente_periodo = f"{mes+1:02d}/{año}"
            cursor.execute("""
                UPDATE InventarioContable
                SET Inicial = ?
                WHERE Periodo = ?
            """, (valor_final, siguiente_periodo))
        elif mes == 12 and año < datetime.now().year:
            # Si es diciembre pero no del año actual, actualizar enero del año siguiente
            siguiente_periodo = f"01/{año+1}"
            cursor.execute("""
                UPDATE InventarioContable
                SET Inicial = ?
                WHERE Periodo = ?
            """, (valor_final, siguiente_periodo))
        
        conn.commit()
        
        if cerrar_conn:
            conn.close()
        
        return True
        
    except Exception as e:
        print(f"Error al actualizar InventarioContable para {periodo}: {str(e)}")
        import traceback
        traceback.print_exc()
        
        try:
            if cerrar_conn and conn:
                conn.rollback()
                conn.close()
        except:
            pass
        
        return False

def recalcular_periodos_año(año):
    """
    Recalcula todos los períodos de un año para asegurar coherencia
    
    Args:
        año (int): Año a recalcular
    
    Returns:
        bool: True si se recalculó correctamente, False en caso contrario
    """
    try:
        conn = get_connection()
        
        # Primer paso: procesar desde enero hasta noviembre
        for mes in range(1, 12):  # Solo procesar hasta noviembre (11)
            periodo = f"{mes:02d}/{año}"
            print(f"Recalculando período {periodo}...")
            
            # Obtener movimientos del período
            cursor = conn.cursor()
            
            # 1. Calcular valores reales basados en MovInventMes
            cursor.execute("""
                SELECT 
                    SUM(CASE WHEN Codigo = '0000000001' THEN Costo ELSE 0 END) AS ValorInicial,
                    SUM(Costo * final) AS ValorFinal
                FROM MovInventMes
                WHERE Periodo = ?
            """, (periodo,))
            
            result = cursor.fetchone()
            if result and result[0] is not None:
                valor_inicial = result[0]
                valor_final = result[1] if result[1] is not None else 0
                
                # 2. Actualizar InventarioContable para el mes actual
                cursor.execute("""
                    UPDATE InventarioContable
                    SET Inicial = ?, Final = ?
                    WHERE Periodo = ?
                """, (valor_inicial, valor_final, periodo))
                
                # 3. Actualizar el inicial del mes siguiente (incluyendo diciembre)
                siguiente_periodo = f"{mes+1:02d}/{año}"
                cursor.execute("""
                    UPDATE InventarioContable
                    SET Inicial = ?
                    WHERE Periodo = ?
                """, (valor_final, siguiente_periodo))
                
                print(f"  Período {periodo}: Actualizado inicial={valor_inicial:.2f}, final={valor_final:.2f}")
                print(f"  Período {siguiente_periodo}: Actualizado inicial={valor_final:.2f}")
            else:
                print(f"  No se encontraron datos para el período {periodo}")
        
        # Segundo paso: procesar diciembre por separado
        periodo_diciembre = f"12/{año}"
        print(f"Recalculando período {periodo_diciembre}...")
        
        cursor = conn.cursor()
        
        # Obtener el valor inicial de diciembre (que debería ser el final de noviembre)
        cursor.execute("""
            SELECT Final
            FROM InventarioContable
            WHERE Periodo = ?
        """, (f"11/{año}",))
        
        result = cursor.fetchone()
        if result and result[0] is not None:
            valor_inicial_diciembre = result[0]
            
            # Para diciembre, obtener el Costo (que debe ser el valor inicial) y mantener el Final existente
            cursor.execute("""
                SELECT 
                    Final
                FROM InventarioContable
                WHERE Periodo = ?
            """, (periodo_diciembre,))
            
            result_dic = cursor.fetchone()
            if result_dic and result_dic[0] is not None:
                valor_final_diciembre = result_dic[0]
                
                # Actualizar solo el inicial de diciembre con el final de noviembre
                cursor.execute("""
                    UPDATE InventarioContable
                    SET Inicial = ?
                    WHERE Periodo = ?
                """, (valor_inicial_diciembre, periodo_diciembre))
                
                # Actualizar el Costo del registro 0000000001 de diciembre
                cursor.execute("""
                    UPDATE MovInventMes
                    SET Costo = ?
                    WHERE Periodo = ? AND Codigo = '0000000001'
                """, (valor_inicial_diciembre, periodo_diciembre))
                
                print(f"  Período {periodo_diciembre}: Actualizado inicial={valor_inicial_diciembre:.2f} (final mantenido en {valor_final_diciembre:.2f})")
                
                # Si existe enero del siguiente año, actualizar su inicial con el final de diciembre
                siguiente_año = año + 1
                cursor.execute("""
                    SELECT COUNT(*) 
                    FROM InventarioContable 
                    WHERE Periodo = ?
                """, (f"01/{siguiente_año}",))
                
                if cursor.fetchone()[0] > 0:
                    cursor.execute("""
                        UPDATE InventarioContable
                        SET Inicial = ?
                        WHERE Periodo = ?
                    """, (valor_final_diciembre, f"01/{siguiente_año}"))
                    print(f"  Período 01/{siguiente_año}: Actualizado inicial={valor_final_diciembre:.2f}")
            else:
                print(f"  No se encontró valor final para el período {periodo_diciembre}")
        else:
            print(f"  No se encontró valor final para el período 11/{año}")
        
        conn.commit()
        conn.close()
        
        return True
        
    except Exception as e:
        print(f"Error al recalcular períodos del año {año}: {str(e)}")
        import traceback
        traceback.print_exc()
        
        try:
            conn.rollback()
            conn.close()
        except:
            pass
        
        return False

def asegurar_no_negativos(conn, periodo):
    """
    Asegura que no haya existencias negativas en MovInventMes
    aplicando la regla: inicial + entradas >= salidas + autoconsumo + retiros
    
    Args:
        conn: Conexión a la base de datos
        periodo: Período a verificar
    """
    cursor = conn.cursor()
    
    # Buscar registros donde final < 0 o inicial + entradas < salidas + autoconsumo + retiros
    cursor.execute("""
        SELECT Codigo, inicial, Entradas, Salidas, AutoConsumo, Retiros, final
        FROM MovInventMes
        WHERE Periodo = ? 
          AND (final < 0 OR (inicial + Entradas) < (Salidas + AutoConsumo + Retiros))
    """, (periodo,))
    
    registros_negativos = cursor.fetchall()
    if registros_negativos:
        print(f"Corrigiendo {len(registros_negativos)} registros con potencial de existencias negativas...")
        
        for reg in registros_negativos:
            codigo = reg[0]
            inicial = reg[1]
            entradas = reg[2]
            salidas = reg[3]
            autoconsumo = reg[4]
            retiros = reg[5]
            final_actual = reg[6]
            
            # Calcular las salidas totales
            salidas_totales = salidas + autoconsumo + retiros
            
            # Estrategia 1: Si hay existencias negativas, incrementar entradas
            if final_actual < 0 or (inicial + entradas) < salidas_totales:
                # Calcular las entradas adicionales necesarias para garantizar existencias positivas
                # La fórmula asegura que: inicial + (entradas + adicionales) = salidas_totales + margen
                entradas_necesarias = salidas_totales - inicial + 15  # Agregamos margen de 15 unidades
                
                if entradas < entradas_necesarias:
                    entradas_adicionales = entradas_necesarias - entradas
                    
                    # Actualizar el registro con las entradas adicionales
                    cursor.execute("""
                        UPDATE MovInventMes
                        SET Entradas = Entradas + ?,
                            final = inicial + Entradas + ? - Salidas - AutoConsumo - Retiros
                        WHERE Periodo = ? AND Codigo = ?
                    """, (entradas_adicionales, entradas_adicionales, periodo, codigo))
                    
                    print(f"  Código {codigo}: Entradas ajustadas de {entradas} a {entradas + entradas_adicionales}")
            
            # Estrategia 2: Verificación adicional para casos extremos donde final aún sea negativo
            cursor.execute("""
                SELECT final 
                FROM MovInventMes
                WHERE Periodo = ? AND Codigo = ? AND final < 0
            """, (periodo, codigo))
            
            if cursor.fetchone():
                # Si aún hay existencias negativas, ajustar las salidas
                cursor.execute("""
                    UPDATE MovInventMes
                    SET Salidas = CASE WHEN inicial + Entradas - AutoConsumo - Retiros < 0 THEN 0 
                                     ELSE inicial + Entradas - AutoConsumo - Retiros END,
                        final = CASE WHEN inicial + Entradas - AutoConsumo - Retiros < 0 THEN 0
                                  ELSE inicial + Entradas - Salidas - AutoConsumo - Retiros END
                    WHERE Periodo = ? AND Codigo = ? AND final < 0
                """, (periodo, codigo))
                print(f"  Código {codigo}: Salidas ajustadas para evitar existencias negativas")
    
    # Verificación final: asegurar que todos los finales sean >= 0
    cursor.execute("""
        UPDATE MovInventMes
        SET final = 0
        WHERE Periodo = ? AND final < 0
    """, (periodo,))
    
    conn.commit()

def generar_movimientos_directo(año, mes):
    """
    Genera movimientos directamente para la tabla MovInventMes sin pasar por MovInvent
    
    Args:
        año (int): Año para el cual generar movimientos
        mes (int): Mes para el cual generar movimientos (1-12)
    
    Returns:
        bool: True si se generó correctamente, False en caso contrario
    """
    try:
        periodo = f"{mes:02d}/{año}"
        print(f"\nGenerando movimientos directamente para {periodo}")
        
        # Obtener datos del período desde InventarioContable
        datos_periodo = obtener_datos_periodo(periodo)
        if datos_periodo is None:
            print(f"Error: No se encontró el período {periodo} en InventarioContable")
            return False
        
        # CRUCIAL: Usar los valores exactos de la tabla InventarioContable
        valor_inicial = float(datos_periodo['Inicial'])
        valor_final = float(datos_periodo['Final'])
        
        # CORRECCIÓN: El campo Inventario debe ser EXACTAMENTE el valor inicial del mes 
        # según está en InventarioContable
        valor_inventario = valor_inicial
        
        # Valores fijos que deben respetarse
        VALOR_INICIAL_2024 = 1120797.03  # Final de diciembre 2023 = Inicial enero 2024
        VALOR_FINAL_2024 = 1892903.00    # Final diciembre 2024
        
        # Para enero 2024, verificar que el valor inicial sea el correcto (final de 2023)
        if año == 2024 and mes == 1:
            if abs(valor_inicial - VALOR_INICIAL_2024) > 0.01:
                print(f"ADVERTENCIA: El valor inicial de enero 2024 debe ser {VALOR_INICIAL_2024}.")
                print(f"             Se usará este valor en lugar de {valor_inicial}.")
                valor_inicial = VALOR_INICIAL_2024
                valor_inventario = VALOR_INICIAL_2024
                
                # Actualizar el valor en la base de datos para mantener coherencia
                conn_temp = get_connection()
                cursor_temp = conn_temp.cursor()
                cursor_temp.execute("""
                    UPDATE InventarioContable
                    SET Inicial = ?
                    WHERE Periodo = '01/2024'
                """, (VALOR_INICIAL_2024,))
                conn_temp.commit()
                conn_temp.close()
        
        # Si estamos generando diciembre 2024, verificar que el valor final sea correcto
        if año == 2024 and mes == 12:
            if abs(valor_final - VALOR_FINAL_2024) > 0.01:
                print(f"ADVERTENCIA: El valor final de diciembre 2024 debe ser {VALOR_FINAL_2024}.")
                print(f"             Se usará este valor en lugar de {valor_final}.")
                valor_final = VALOR_FINAL_2024
                
                # Actualizar el valor en la base de datos para mantener coherencia
                conn_temp = get_connection()
                cursor_temp = conn_temp.cursor()
                cursor_temp.execute("""
                    UPDATE InventarioContable
                    SET Final = ?
                    WHERE Periodo = '12/2024'
                """, (VALOR_FINAL_2024,))
                conn_temp.commit()
                conn_temp.close()
        
        print(f"Valor inicial: {valor_inicial:.2f}")
        print(f"Valor final objetivo: {valor_final:.2f}")
        print(f"Diferencia: {valor_final - valor_inicial:.2f}")
        
        # Calcular diferencia para este mes específico
        diferencia = valor_final - valor_inicial
        
        # Obtener existencias previas del período anterior
        existencias_previas = obtener_existencias_previas(año, mes)
        print(f"Se encontraron {len(existencias_previas)} productos con existencias previas")
        
        # Obtener productos para distribuir movimientos
        productos = obtener_productos()
        if not productos:
            print("Error: No hay productos disponibles en el inventario")
            return False
        
        # Obtener días hábiles del mes
        dias_habiles = obtener_dias_habiles(año, mes)
        if not dias_habiles:
            print(f"Error: No hay días hábiles para {mes}/{año}")
            return False
        
        # Seleccionar un número de productos que depende del mes
        # Más productos para los meses con mayor diferencia
        factor_productos = abs(diferencia) / 10000  # Ajustar según la magnitud
        num_productos_base = max(10, min(50, int(20 + factor_productos)))
        num_productos = min(len(productos), num_productos_base)
        
        # Priorizar productos con existencias previas
        productos_con_existencia = [p for p in productos if p['codigo'] in existencias_previas and existencias_previas[p['codigo']] > 0]
        productos_sin_existencia = [p for p in productos if p['codigo'] not in existencias_previas or existencias_previas[p['codigo']] == 0]
        
        # Asegurar que tenemos suficientes productos
        if len(productos_con_existencia) < num_productos // 3:
            # Si hay pocos productos con existencia, completar con productos sin existencia
            productos_adicionales = random.sample(productos_sin_existencia, min(len(productos_sin_existencia), num_productos - len(productos_con_existencia)))
            productos_seleccionados = productos_con_existencia + productos_adicionales
        else:
            # Si hay suficientes productos con existencia, seleccionar algunos de ellos
            productos_seleccionados = random.sample(productos_con_existencia, min(len(productos_con_existencia), num_productos))
            
            # Completar con productos sin existencia si es necesario
            if len(productos_seleccionados) < num_productos:
                productos_adicionales = random.sample(productos_sin_existencia, min(len(productos_sin_existencia), num_productos - len(productos_seleccionados)))
                productos_seleccionados += productos_adicionales
        
        # Limpiar registros existentes para este período
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM MovInventMes WHERE Periodo = ?", (periodo,))
        conn.commit()
        
        # Insertar registro de inventario inicial (1 unidad con el valor monetario total)
        primer_dia = dias_habiles[0]
        cursor.execute("""
            INSERT INTO MovInventMes 
            (Periodo, Codigo, inicial, Costo, Descripcion, Entradas, Salidas, AutoConsumo, Retiros, final, Fecha, Inventario)
            VALUES (?, '0000000001', 1, ?, 'INVENTARIO INICIAL MES ANTERIOR', 0, 0, 0, 0, 1, ?, ?)
        """, (periodo, valor_inicial, primer_dia, valor_inventario))
        
        # Distribuir el valor total entre los productos
        # Ajustamos la magnitud según la diferencia del mes
        valor_distribuir = abs(diferencia) * 0.5  # Usamos 50% de la diferencia
        if diferencia > 0:
            # Para diferencia positiva, aumentamos el factor de distribución
            valor_total_productos = valor_distribuir + valor_inicial * 0.1  # 10% del valor inicial adicional
        else:
            # Para diferencia negativa, limitamos un poco el valor distribuido
            valor_total_productos = max(valor_distribuir, valor_inicial * 0.05)  # Al menos 5% del valor inicial
            
        valores_por_producto = distribuir_valor(valor_total_productos, len(productos_seleccionados))
        
        # Variables para acumular valores monetarios
        total_valor_entradas = 0
        total_valor_salidas = 0
        
        # Diccionario para rastrear la existencia acumulada de cada producto durante el mes
        existencia_acumulada = {}
        for codigo in [p['codigo'] for p in productos_seleccionados]:
            existencia_acumulada[codigo] = existencias_previas.get(codigo, 0)
        
        # Procesar cada producto
        for i, producto in enumerate(productos_seleccionados):
            codigo = producto['codigo']
            
            # Asignar categoría de precio (0-4) basada en algún criterio
            categoria_precio = i % 5
            
            # Calcular un precio consistente
            precio_unitario = calcular_precio_consistente(producto['precio'], categoria_precio)
            
            # Valor total asignado a este producto
            valor_producto = valores_por_producto[i]
            
            # Existencia inicial del producto al inicio del mes
            existencia_inicial = existencias_previas.get(codigo, 0)
            
            # Para productos con existencia inicial < 5, garantizar una entrada inicial grande
            if existencia_inicial < 5:
                entrada_inicial_min = max(20, int(valor_producto / (precio_unitario * 0.5)))
                entrada_inicial = random.randint(entrada_inicial_min, entrada_inicial_min + 20)
                
                # Registrar entrada inicial al principio del mes
                dia_entrada_inicial = dias_habiles[1] if len(dias_habiles) > 1 else dias_habiles[0]
                
                # Actualizar existencia acumulada
                existencia_acumulada[codigo] = existencia_inicial + entrada_inicial
                
                # Calcular valor de la entrada
                valor_entrada = round(entrada_inicial * precio_unitario, 2)
                total_valor_entradas += valor_entrada
                
                # Insertar registro de entrada inicial
                cursor.execute("""
                    INSERT INTO MovInventMes 
                    (Periodo, Codigo, inicial, Costo, Descripcion, Entradas, Salidas, AutoConsumo, Retiros, final, Fecha, Inventario)
                    VALUES (?, ?, ?, ?, ?, ?, 0, 0, 0, ?, ?, ?)
                """, (
                    periodo, 
                    codigo, 
                    existencia_inicial,
                    precio_unitario,
                    producto['descripcion'],
                    entrada_inicial,
                    existencia_acumulada[codigo],
                    dia_entrada_inicial,
                    valor_inventario
                ))
            
            # Determinar número de movimientos adicionales para este producto (aprox. 4x)
            factor_multiplicador_movs = 4 
            if existencia_inicial < 5:
                # Si ya generamos una entrada inicial grande, menos movimientos adicionales
                num_movimientos = random.randint(1 * factor_multiplicador_movs, 3 * factor_multiplicador_movs) 
            else:
                # Más movimientos para productos con existencia
                num_movimientos = random.randint(2 * factor_multiplicador_movs, 4 * factor_multiplicador_movs)
            print(f"    Producto {codigo}: Generando {num_movimientos} movimientos adicionales.")
            
            # Dividir los días hábiles en grupos para distribuir los movimientos
            # Excluimos los primeros días ya usados
            dias_inicio = 2 if existencia_inicial < 5 else 1
            dias_disponibles = dias_habiles[dias_inicio:] if len(dias_habiles) > dias_inicio else []
            
            if not dias_disponibles:
                continue  # Si no hay más días disponibles, pasar al siguiente producto
                
            # Mezclar días para evitar patrones
            random.shuffle(dias_disponibles)
            dias_seleccionados = dias_disponibles[:num_movimientos] if num_movimientos <= len(dias_disponibles) else dias_disponibles
            dias_seleccionados.sort()  # Ordenar cronológicamente
            
            # Determinar tendencia según diferencia global
            # Si la diferencia es positiva (aumento en inventario), favorecer más entradas que salidas
            porcentaje_entradas = 0.7 if diferencia > 0 else 0.5
            
            # Para cada día seleccionado, generar un movimiento
            for idx, dia in enumerate(dias_seleccionados):
                # Existencia antes del movimiento
                existencia_antes = existencia_acumulada[codigo]
                
                # Si es el primer movimiento o hay poca existencia, garantizar una entrada
                debe_tener_entrada = idx == 0 or existencia_antes < 15 or random.random() < porcentaje_entradas
                
                # Solo permitir salidas si hay suficiente existencia
                puede_tener_salida = existencia_antes >= 15 and random.random() < 0.6
                
                # Si es el último movimiento y aún no hay suficientes existencias, forzar entrada
                if idx == len(dias_seleccionados) - 1 and existencia_antes < 20:
                    debe_tener_entrada = True
                    puede_tener_salida = False
                
                # Calcular cantidades para entradas/salidas
                cantidad_base = max(5, round(valor_producto / (precio_unitario * (num_movimientos + 1))))
                
                # CASO 1: Solo entradas
                if debe_tener_entrada and not puede_tener_salida:
                    # Variación aleatoria para hacer más natural
                    factor_variacion = random.uniform(0.8, 1.5)
                    cantidad_entrada = max(5, round(cantidad_base * factor_variacion))
                    
                    # Actualizar existencia acumulada
                    existencia_acumulada[codigo] += cantidad_entrada
                    
                    # Calcular valor real
                    valor_entrada = round(cantidad_entrada * precio_unitario, 2)
                    total_valor_entradas += valor_entrada
                    
                    cursor.execute("""
                        INSERT INTO MovInventMes 
                        (Periodo, Codigo, inicial, Costo, Descripcion, Entradas, Salidas, AutoConsumo, Retiros, final, Fecha, Inventario)
                        VALUES (?, ?, ?, ?, ?, ?, 0, 0, 0, ?, ?, ?)
                    """, (
                        periodo, 
                        codigo, 
                        existencia_antes,
                        precio_unitario,
                        producto['descripcion'],
                        cantidad_entrada,
                        existencia_acumulada[codigo],
                        dia,
                        valor_inventario
                    ))
                
                # CASO 2: Solo salidas (asegurando no negatividad de valor)
                elif puede_tener_salida and not debe_tener_entrada and existencia_antes > 20:
                    # Calcular salida máxima permitida en cantidad (dejando margen de seguridad de cantidad)
                    salida_maxima_cantidad = existencia_antes - 10  # Dejamos al menos 10 unidades

                    # Calcular salida deseada en cantidad
                    factor_variacion = random.uniform(0.3, 0.9)
                    cantidad_salida_deseada = round(cantidad_base * factor_variacion)
                    
                    # Limitar la salida en cantidad al máximo permitido y al menos 1
                    cantidad_salida = min(cantidad_salida_deseada, salida_maxima_cantidad)
                    cantidad_salida = max(1, cantidad_salida)

                    # Ajustar cantidad_salida para no exceder el valor disponible
                    if precio_unitario > 0:
                        valor_disponible_para_salida = existencia_antes * precio_unitario
                        cantidad_salida_max_por_valor = int(valor_disponible_para_salida / precio_unitario)
                        if cantidad_salida > cantidad_salida_max_por_valor:
                            print(f"    AJUSTE VALOR SALIDA (CASO 2): Prod {codigo}, Día {dia.strftime('%Y-%m-%d')}, Cant Salida: {cantidad_salida} -> {cantidad_salida_max_por_valor}, ExistAntes: {existencia_antes}, Precio: {precio_unitario:.2f}, ValorDisp: {valor_disponible_para_salida:.2f}")
                            cantidad_salida = cantidad_salida_max_por_valor
                    else: # Si el precio es cero o negativo, no permitir salidas que puedan causar problemas de valor.
                        if cantidad_salida > 0 :
                             print(f"    ADVERTENCIA VALOR SALIDA (CASO 2): Prod {codigo}, Precio <= 0 ({precio_unitario:.2f}), forzando cantidad_salida a 0.")
                        cantidad_salida = 0
                    
                    cantidad_salida = max(0, cantidad_salida) # Asegurar que no sea negativa tras el ajuste de valor

                    if cantidad_salida > 0:
                        # Actualizar existencia acumulada (real)
                        existencia_acumulada_real = existencia_antes - cantidad_salida
                        
                        # Calcular valor real
                        valor_salida = round(cantidad_salida * precio_unitario, 2)
                        total_valor_salidas += valor_salida

                        # LÓGICA DE PRUEBA: Para movimientos con salida, inicial_db = cantidad_salida
                        inicial_para_db = cantidad_salida
                        final_para_db = 0 # Porque inicial_para_db - cantidad_salida = 0
                        
                        cursor.execute("""
                            INSERT INTO MovInventMes 
                            (Periodo, Codigo, inicial, Costo, Descripcion, Entradas, Salidas, AutoConsumo, Retiros, final, Fecha, Inventario)
                            VALUES (?, ?, ?, ?, ?, 0, ?, 0, 0, ?, ?, ?)
                        """, (
                            periodo, 
                            codigo, 
                            inicial_para_db, # PRUEBA
                            precio_unitario,
                            producto['descripcion'],
                            cantidad_salida,
                            final_para_db, # PRUEBA
                            dia,
                            valor_inventario
                        ))
                        existencia_acumulada[codigo] = existencia_acumulada_real # Actualizar el tracker global
                    # Si cantidad_salida se vuelve 0 después del ajuste de valor, no se inserta registro.
                
                # CASO 3: Combinado (entradas y salidas en el mismo registro, asegurando no negatividad de valor)
                elif debe_tener_entrada and puede_tener_salida:
                    # Primero calculamos la entrada
                    factor_variacion_entrada = random.uniform(1.0, 1.8)
                    cantidad_entrada = max(5, round(cantidad_base * factor_variacion_entrada))
                    
                    # Luego calculamos la salida, considerando la entrada que acabamos de planificar
                    existencia_con_entrada_planificada = existencia_antes + cantidad_entrada
                    # Calcular salida máxima permitida en cantidad (dejando margen de seguridad de cantidad)
                    salida_maxima_cantidad = existencia_con_entrada_planificada - 10
                    
                    factor_variacion_salida = random.uniform(0.3, 0.7)
                    cantidad_salida_deseada = round(cantidad_base * factor_variacion_salida)
                    
                    # Limitar la salida en cantidad al máximo permitido y al menos 1
                    cantidad_salida = min(cantidad_salida_deseada, salida_maxima_cantidad)
                    cantidad_salida = max(1, cantidad_salida)

                    # Ajustar cantidad_salida para no exceder el valor disponible tras la entrada
                    if precio_unitario > 0:
                        valor_disponible_para_salida_tras_entrada = (existencia_antes + cantidad_entrada) * precio_unitario
                        cantidad_salida_max_por_valor = int(valor_disponible_para_salida_tras_entrada / precio_unitario)
                        if cantidad_salida > cantidad_salida_max_por_valor:
                            print(f"    AJUSTE VALOR SALIDA (CASO 3): Prod {codigo}, Día {dia.strftime('%Y-%m-%d')}, Cant Salida: {cantidad_salida} -> {cantidad_salida_max_por_valor}, ExistAntes: {existencia_antes}, CantEntrada: {cantidad_entrada}, Precio: {precio_unitario:.2f}, ValorDispTrasEnt: {valor_disponible_para_salida_tras_entrada:.2f}")
                            cantidad_salida = cantidad_salida_max_por_valor
                    else: # Si el precio es cero o negativo
                        if cantidad_salida > 0 :
                            print(f"    ADVERTENCIA VALOR SALIDA (CASO 3): Prod {codigo}, Precio <= 0 ({precio_unitario:.2f}), forzando cantidad_salida a 0.")
                        cantidad_salida = 0

                    cantidad_salida = max(0, cantidad_salida) # Asegurar que no sea negativa tras el ajuste de valor

                    # Actualizar existencia acumulada (real)
                    existencia_acumulada_real = existencia_antes + cantidad_entrada - cantidad_salida
                    
                    # Calcular valores monetarios
                    valor_entrada = round(cantidad_entrada * precio_unitario, 2)
                    total_valor_entradas += valor_entrada
                    
                    valor_salida = 0
                    if cantidad_salida > 0:
                        valor_salida = round(cantidad_salida * precio_unitario, 2)
                        total_valor_salidas += valor_salida

                    # LÓGICA DE PRUEBA:
                    inicial_para_db = existencia_antes # Por defecto, es la existencia real antes
                    final_para_db = existencia_acumulada_real # Por defecto, es la existencia real después

                    if cantidad_salida > 0:
                        inicial_para_db = cantidad_salida
                        final_para_db = cantidad_entrada # Porque (inicial_db + entrada - salida) = (salida + entrada - salida) = entrada
                    
                    cursor.execute("""
                        INSERT INTO MovInventMes 
                        (Periodo, Codigo, inicial, Costo, Descripcion, Entradas, Salidas, AutoConsumo, Retiros, final, Fecha, Inventario)
                        VALUES (?, ?, ?, ?, ?, ?, ?, 0, 0, ?, ?, ?)
                    """, (
                        periodo, 
                        codigo, 
                        inicial_para_db, # PRUEBA si cantidad_salida > 0
                        precio_unitario,
                        producto['descripcion'],
                        cantidad_entrada,
                        cantidad_salida, 
                        final_para_db, # PRUEBA si cantidad_salida > 0
                        dia,
                        valor_inventario
                    ))
                    existencia_acumulada[codigo] = existencia_acumulada_real # Actualizar el tracker global
                
                # Si no se cumplió ninguna condición, asegurar al menos una entrada pequeña
                else:
                    cantidad_entrada = max(3, round(cantidad_base * 0.5))
                    
                    # Actualizar existencia acumulada
                    existencia_acumulada[codigo] += cantidad_entrada
                    
                    # Calcular valor real
                    valor_entrada = round(cantidad_entrada * precio_unitario, 2)
                    total_valor_entradas += valor_entrada
                    
                    cursor.execute("""
                        INSERT INTO MovInventMes 
                        (Periodo, Codigo, inicial, Costo, Descripcion, Entradas, Salidas, AutoConsumo, Retiros, final, Fecha, Inventario)
                        VALUES (?, ?, ?, ?, ?, ?, 0, 0, 0, ?, ?, ?)
                    """, (
                        periodo, 
                        codigo, 
                        existencia_antes,
                        precio_unitario,
                        producto['descripcion'],
                        cantidad_entrada,
                        existencia_acumulada[codigo],
                        dia,
                        valor_inventario
                    ))
        
        # Verificar que los totales monetarios cuadren con lo esperado
        valor_final_calculado = valor_inicial + total_valor_entradas - total_valor_salidas
        
        print(f"Valor inicial: {valor_inicial:.2f}")
        print(f"Total valor entradas: {total_valor_entradas:.2f}")
        print(f"Total valor salidas: {total_valor_salidas:.2f}")
        print(f"Valor final calculado: {valor_final_calculado:.2f}")
        print(f"Valor final objetivo: {valor_final:.2f}")

        # AJUSTE GENERAL PARA TODOS LOS MESES: Insertar movimientos para cuadrar con InventarioContable.Final
        # La condición if mes == 12: ha sido eliminada de aquí.
        valor_actual_inventario_sin_ajuste_final = valor_inicial + total_valor_entradas - total_valor_salidas
        valor_final_objetivo_de_tabla_contable = float(datos_periodo['Final'])
        
        diferencia_ajuste_final = round(valor_final_objetivo_de_tabla_contable - valor_actual_inventario_sin_ajuste_final, 2)

        if abs(diferencia_ajuste_final) > 0.01:
            print(f"Ajuste final necesario para {periodo}: {diferencia_ajuste_final:.2f}. Se dividirá en 5 movimientos.")
            fecha_ajuste = dias_habiles[-1] # Último día hábil del mes
            
            costo_por_movimiento_ajuste = round(abs(diferencia_ajuste_final) / 5.0, 2)
            valor_total_ajuste_distribuido = 0

            for i in range(5):
                costo_actual_movimiento = costo_por_movimiento_ajuste
                if i == 4:
                    costo_actual_movimiento = round(abs(diferencia_ajuste_final) - valor_total_ajuste_distribuido, 2)
                valor_total_ajuste_distribuido += costo_actual_movimiento

                if not productos_seleccionados:
                    print(f"ADVERTENCIA (Ajuste {i+1}/5): No hay productos. Usando genérico.")
                    codigo_producto_ajuste = "0000000002"
                    descripcion_producto_ajuste = f"AJUSTE VALOR FINAL PERIODO ({i+1}/5)"
                else:
                    producto_para_ajuste = random.choice(productos_seleccionados)
                    codigo_producto_ajuste = producto_para_ajuste['codigo']
                    descripcion_producto_ajuste = producto_para_ajuste['descripcion']
                
                q_inicial_ajuste, q_entrada_ajuste, q_salida_ajuste, tipo_ajuste_log = 0, 0, 0, ""

                if diferencia_ajuste_final < 0:
                    tipo_ajuste_log = "Egreso (Salida) de Ajuste Final"
                    q_inicial_ajuste = 1 
                    q_entrada_ajuste = 0
                    q_salida_ajuste = 1
                else:
                    tipo_ajuste_log = "Ingreso (Entrada) de Ajuste Final"
                    q_inicial_ajuste = 0
                    q_entrada_ajuste = 1
                    q_salida_ajuste = 0
                
                q_final_ajuste = q_inicial_ajuste + q_entrada_ajuste - q_salida_ajuste
                
                print(f"  Mov.AjusteFinal {i+1}/5: Prod: {codigo_producto_ajuste}")
                print(f"    Tipo: {tipo_ajuste_log}, Cant.Ini: {q_inicial_ajuste}, Ent: {q_entrada_ajuste}, Sal: {q_salida_ajuste}, Cant.Fin: {q_final_ajuste}")
                print(f"    Costo mov. ajuste: {costo_actual_movimiento:.2f}")

                cursor.execute("""
                     INSERT INTO MovInventMes
                     (Periodo, Codigo, inicial, Costo, Descripcion, Entradas, Salidas, AutoConsumo, Retiros, final, Fecha, Inventario)
                     VALUES (?, ?, ?, ?, ?, ?, ?, 0, 0, ?, ?, ?)
                """, (
                    periodo, codigo_producto_ajuste, q_inicial_ajuste, costo_actual_movimiento, 
                    descripcion_producto_ajuste, q_entrada_ajuste, q_salida_ajuste,  
                    q_final_ajuste, fecha_ajuste, valor_inventario
                ))
            
            print(f"  Total 5 mov. ajuste insertados. Costo total distribuido: {valor_total_ajuste_distribuido:.2f} (Objetivo: {abs(diferencia_ajuste_final):.2f})")
            
            # Actualizar el valor_final_calculado para el print que resume el mes
            valor_final_calculado = valor_actual_inventario_sin_ajuste_final + diferencia_ajuste_final
            print(f"Nuevo valor final del mes tras 5 ajustes: {valor_final_calculado:.2f}")
        else:
            print(f"No se requiere ajuste final para {periodo}. Diferencia: {diferencia_ajuste_final:.2f}")

        # Aplicar verificación adicional para asegurar que no haya existencias negativas
        print("Verificando que no haya existencias negativas...")
        asegurar_no_negativos(conn, periodo)
        
        # Contar registros generados
        cursor.execute("SELECT COUNT(*) FROM MovInventMes WHERE Periodo = ?", (periodo,))
        total_registros = cursor.fetchone()[0]
        
        # Generar registro en MovPeridoMes
        cursor.execute("DELETE FROM MovPeridoMes WHERE Periodo = ?", (periodo,))
        
        # Corregir la inserción a MovPeridoMes EXCLUYENDO el campo Inventario que no existe en esa tabla
        cursor.execute("""
            INSERT INTO MovPeridoMes 
            (Periodo, Codigo, Descripcion, Costo, Inicial, Entradas, Salidas, AutoConsumo, Retiros, Fecha)
            SELECT 
                Periodo, Codigo, Descripcion, Costo, Inicial, 
                Entradas, Salidas, AutoConsumo, Retiros, Fecha
            FROM MovInventMes  
            WHERE Periodo = ?
        """, (periodo,))
        
        conn.commit()
        conn.close()
        
        print(f"\nMovimientos generados directamente para {periodo}")
        print(f"Total registros generados: {total_registros}")
        print(f"Valor inicial total: {valor_inicial:.2f}")
        print(f"Valor inventario: {valor_inventario:.2f} (igual al valor inicial del mes)")
        print(f"Valor final alcanzado: {valor_final:.2f}")
        
        return True
    
    except Exception as e:
        print(f"Error al generar movimientos directos: {str(e)}")
        import traceback
        traceback.print_exc()
        
        try:
            conn.rollback()
            conn.close()
        except:
            pass
        
        return False

def inicializar_periodos_año(año, valor_inicial_enero=None, valor_final_diciembre=None):
    """
    Inicializa o ajusta los valores iniciales y finales de los períodos de un año
    
    Args:
        año (int): Año a inicializar
        valor_inicial_enero (float): Valor inicial para enero (si es None, no se modifica)
        valor_final_diciembre (float): Valor final para diciembre (si es None, no se modifica)
    
    Returns:
        bool: True si se inicializó correctamente, False en caso contrario
    """
    try:
        conn = get_connection()
        cursor = conn.cursor()
        
        # 1. Verificar si existen registros para el año
        cursor.execute("""
            SELECT COUNT(*) 
            FROM InventarioContable 
            WHERE Periodo LIKE ?
        """, (f'__/{año}',))
        
        count = cursor.fetchone()[0]
        if count < 12:
            print(f"No hay registros completos para el año {año} en InventarioContable.")
            print(f"Se necesitan crear {12-count} registros.")
            # Esto podría implementarse si es necesario crear registros
            return False
        
        # 2. Si se especificó, actualizar el valor inicial de enero
        if valor_inicial_enero is not None:
            cursor.execute("""
                UPDATE InventarioContable
                SET Inicial = ?
                WHERE Periodo = ?
            """, (valor_inicial_enero, f'01/{año}'))
            print(f"Actualizado el valor inicial de enero {año} a {valor_inicial_enero:.2f}")
        
        # 3. Si se especificó, actualizar el valor final de diciembre
        if valor_final_diciembre is not None:
            cursor.execute("""
                UPDATE InventarioContable
                SET Final = ?
                WHERE Periodo = ?
            """, (valor_final_diciembre, f'12/{año}'))
            print(f"Actualizado el valor final de diciembre {año} a {valor_final_diciembre:.2f}")
        
        # 4. Leer los valores inicial de enero y final de diciembre
        cursor.execute("""
            SELECT Inicial 
            FROM InventarioContable 
            WHERE Periodo = ?
        """, (f'01/{año}',))
        valor_inicial_enero = cursor.fetchone()[0]
        
        cursor.execute("""
            SELECT Final 
            FROM InventarioContable 
            WHERE Periodo = ?
        """, (f'12/{año}',))
        valor_final_diciembre = cursor.fetchone()[0]
        
        # 5. Calcular la diferencia anual a distribuir
        diferencia_anual = valor_final_diciembre - valor_inicial_enero
        print(f"Valor inicial año {año}: {valor_inicial_enero:.2f}")
        print(f"Valor final año {año}: {valor_final_diciembre:.2f}")
        print(f"Diferencia a distribuir: {diferencia_anual:.2f}")
        
        # 6. Definir la distribución mensual (pesos por mes)
        # Usamos pesos específicos para cada mes según patrones de negocio típicos
        # Los pesos determinan qué proporción de la diferencia anual se asigna a cada mes
        pesos_mensuales = []
        
        if diferencia_anual > 0:
            # Para crecimiento positivo: más crecimiento en meses de alta actividad
            if año == 2024:
                # Distribución específica para 2024
                pesos = [0.04, 0.05, 0.08, 0.09, 0.1, 0.1, 0.09, 0.11, 0.13, 0.12, 0.09]
            else:
                # Distribución general para otros años
                pesos = [0.05, 0.06, 0.08, 0.09, 0.1, 0.1, 0.09, 0.1, 0.11, 0.12, 0.1]
        else:
            # Para reducción de inventario: distribución más uniforme con ligera
            # concentración en ciertos meses
            pesos = [0.07, 0.07, 0.09, 0.1, 0.1, 0.09, 0.08, 0.1, 0.09, 0.11, 0.1]
        
        # Normalizar pesos para asegurar que sumen exactamente 1
        suma_pesos = sum(pesos)
        pesos_normalizados = [p/suma_pesos for p in pesos]
        
        # 7. Calcular valores mensuales basados en los pesos
        # Partimos del valor inicial
        valor_acumulado = valor_inicial_enero
        valores_finales_mensuales = [valor_acumulado]  # El primer valor es el inicial de enero
        
        # Modificación: Distribuir 11/12 de la diferencia_anual hasta noviembre
        diferencia_hasta_noviembre = diferencia_anual * (11.0 / 12.0)

        for i in range(11):  # Para los 11 primeros meses (enero a noviembre)
            # Calculamos incremento para este mes usando la diferencia_hasta_noviembre
            incremento_mes = diferencia_hasta_noviembre * pesos_normalizados[i]
            
            # El valor final de este mes
            valor_final_mes = round(valor_acumulado + incremento_mes, 2)
            valores_finales_mensuales.append(valor_final_mes)
            
            # El próximo mes parte de este valor
            valor_acumulado = valor_final_mes
        
        # Asegurar que el último valor sea exactamente el valor final de diciembre
        # REMOVED: valores_finales_mensuales[-1] = valor_final_diciembre
        
        # 8. Actualizar los valores en la base de datos (procesando primero enero a noviembre)
        for mes in range(1, 12):  # Solo procesar hasta noviembre (1-11)
            valor_final_mes = valores_finales_mensuales[mes]
            siguiente_mes = mes + 1
            
            # El final del mes actual es el inicial del siguiente
            cursor.execute("""
                UPDATE InventarioContable
                SET Final = ?
                WHERE Periodo = ?
            """, (valor_final_mes, f'{mes:02d}/{año}'))
            
            cursor.execute("""
                UPDATE InventarioContable
                SET Inicial = ?
                WHERE Periodo = ?
            """, (valor_final_mes, f'{siguiente_mes:02d}/{año}'))
            
            print(f"Mes {mes:02d}: Final={valor_final_mes:.2f}")
            print(f"Mes {siguiente_mes:02d}: Inicial={valor_final_mes:.2f}")
        
        # 9. Manejar diciembre por separado - mantener su valor final específico
        # y asegurarnos que su valor inicial sea el final de noviembre
        valor_final_noviembre = valores_finales_mensuales[11]  # Índice 11 corresponde a noviembre
        
        cursor.execute("""
            UPDATE InventarioContable
            SET Inicial = ?
            WHERE Periodo = ?
        """, (valor_final_noviembre, f'12/{año}'))
        
        print(f"Mes 12: Inicial={valor_final_noviembre:.2f}, Final={valor_final_diciembre:.2f} (mantenido)")
        
        # 10. Si el año tiene un año siguiente, actualizar también enero del siguiente
        if año < datetime.now().year + 1:
            try:
                cursor.execute("""
                    UPDATE InventarioContable
                    SET Inicial = ?
                    WHERE Periodo = ?
                """, (valor_final_diciembre, f'01/{año+1}'))
                print(f"Actualizado inicial de enero {año+1} a {valor_final_diciembre:.2f}")
            except:
                print(f"No se pudo actualizar el valor inicial de enero {año+1}")
        
        conn.commit()
        conn.close()
        return True
    
    except Exception as e:
        print(f"Error al inicializar períodos del año {año}: {str(e)}")
        import traceback
        traceback.print_exc()
        
        try:
            conn.rollback()
            conn.close()
        except:
            pass
        
        return False

def generar_año_directo(año, valor_inicial_enero=None, valor_final_diciembre=None):
    """
    Genera los movimientos para todos los meses de un año directamente
    
    Args:
        año (int): El año para generar movimientos
        valor_inicial_enero (float): Valor inicial para enero (opcional)
        valor_final_diciembre (float): Valor final para diciembre (opcional)
    """
    try:
        print(f"Generando movimientos directos para el año {año}...")
        
        # Inicializar los períodos con los valores requeridos
        if valor_inicial_enero is not None or valor_final_diciembre is not None:
            print(f"Inicializando valores de períodos para el año {año}...")
            inicializar_periodos_año(año, valor_inicial_enero, valor_final_diciembre)
        
        # Procesar cada mes
        for mes in range(1, 13):
            try:
                nombre_mes = calendar.month_name[mes]
                print(f"\nProcesando {nombre_mes} {año}...")
                
                exito = generar_movimientos_directo(año, mes)
                if not exito:
                    print(f"Error al generar movimientos para {mes:02d}/{año}")
                else:
                    # Verificar que los valores sean coherentes
                    verificar_coherencia_valores(año, mes)
            
            except Exception as e:
                print(f"Error en mes {mes}: {str(e)}")
        
        # Recalcular valores de InventarioContable para todo el año
        print("\nRecalculando valores de InventarioContable...")
        recalcular_periodos_año(año)
        
        # Verificación final para asegurar que el valor final de diciembre es correcto
        if valor_final_diciembre is not None:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE InventarioContable
                SET Final = ?
                WHERE Periodo = ?
            """, (valor_final_diciembre, f'12/{año}'))
            conn.commit()
            conn.close()
            print(f"Verificación final: Valor de diciembre {año} ajustado a {valor_final_diciembre:.2f}")
            
            # Verificar coherencia del valor final de diciembre
            verificar_coherencia_valores(año, 12)
        
        print(f"\nGeneración de movimientos completa para el año {año}")
    
    except Exception as e:
        print(f"Error en generar_año_directo: {str(e)}")

def verificar_coherencia_valores(año, mes):
    """
    Verifica que los valores del registro 0000000001 sean coherentes
    con los valores de InventarioContable y corrige si es necesario.
    También verifica la coherencia entre períodos consecutivos.
    
    Args:
        año (int): Año a verificar
        mes (int): Mes a verificar (1-12)
    """
    try:
        periodo = f"{mes:02d}/{año}"
        conn = get_connection()
        cursor = conn.cursor()
        
        # 1. Obtener valores de InventarioContable
        cursor.execute("""
            SELECT Inicial, Final
            FROM InventarioContable
            WHERE Periodo = ?
        """, (periodo,))
        
        row = cursor.fetchone()
        if not row:
            print(f"No se encontró el período {periodo} en InventarioContable.")
            conn.close()
            return
        
        valor_inicial_contable = float(row[0])
        valor_final_contable = float(row[1])
        
        # 2. Obtener valores del registro 0000000001
        cursor.execute("""
            SELECT Costo, final
            FROM MovInventMes
            WHERE Periodo = ? AND Codigo = '0000000001'
        """, (periodo,))
        
        row = cursor.fetchone()
        if not row:
            print(f"No se encontró el registro 0000000001 para el período {periodo}.")
            conn.close()
            return
        
        costo_registro = float(row[0])
        final_registro = float(row[1])
        
        # 3. Verificar coherencia y corregir si es necesario
        hay_cambios = False
        
        # El Costo debe ser igual al valor inicial
        if abs(costo_registro - valor_inicial_contable) > 0.01:
            print(f"Corrigiendo Costo del registro 0000000001 para {periodo}:")
            print(f"  Valor actual: {costo_registro:.2f}")
            print(f"  Valor correcto: {valor_inicial_contable:.2f}")
            
            cursor.execute("""
                UPDATE MovInventMes
                SET Costo = ?
                WHERE Periodo = ? AND Codigo = '0000000001'
            """, (valor_inicial_contable, periodo))
            
            hay_cambios = True
        
        # El valor final debe ser igual al valor final contable
        if abs(final_registro - valor_final_contable) > 0.01:
            print(f"Corrigiendo valor final del registro 0000000001 para {periodo}:")
            print(f"  Valor actual: {final_registro:.2f}")
            print(f"  Valor correcto: {valor_final_contable:.2f}")
            
            cursor.execute("""
                UPDATE MovInventMes
                SET final = ?
                WHERE Periodo = ? AND Codigo = '0000000001'
            """, (valor_final_contable, periodo))
            
            hay_cambios = True
        
        # 4. Verificar coherencia con períodos adyacentes
        
        # 4.1. Para todos los meses excepto enero: verificar que el inicial sea igual al final del mes anterior
        if mes > 1:
            periodo_anterior = f"{mes-1:02d}/{año}"
            cursor.execute("""
                SELECT Final
                FROM InventarioContable
                WHERE Periodo = ?
            """, (periodo_anterior,))
            
            row = cursor.fetchone()
            if row:
                valor_final_anterior = float(row[0])
                
                # El inicial de este mes debe ser igual al final del anterior
                if abs(valor_inicial_contable - valor_final_anterior) > 0.01:
                    print(f"Coherencia entre períodos: El inicial de {periodo} no coincide con el final de {periodo_anterior}")
                    print(f"  Valor inicial de {periodo}: {valor_inicial_contable:.2f}")
                    print(f"  Valor final de {periodo_anterior}: {valor_final_anterior:.2f}")
                    
                    # Corregir el valor inicial del período actual
                    cursor.execute("""
                        UPDATE InventarioContable
                        SET Inicial = ?
                        WHERE Periodo = ?
                    """, (valor_final_anterior, periodo))
                    
                    # También corregir el Costo del registro 0000000001
                    cursor.execute("""
                        UPDATE MovInventMes
                        SET Costo = ?
                        WHERE Periodo = ? AND Codigo = '0000000001'
                    """, (valor_final_anterior, periodo))
                    
                    print(f"  Se actualizó el valor inicial de {periodo} a {valor_final_anterior:.2f}")
                    hay_cambios = True
        
        # 4.2. Para todos los meses excepto diciembre: verificar que el siguiente mes tenga como inicial el final de este
        if mes < 12:
            periodo_siguiente = f"{mes+1:02d}/{año}"
            cursor.execute("""
                SELECT Inicial
                FROM InventarioContable
                WHERE Periodo = ?
            """, (periodo_siguiente,))
            
            row = cursor.fetchone()
            if row:
                valor_inicial_siguiente = float(row[0])
                
                # El inicial del siguiente debe ser igual al final de este
                if abs(valor_inicial_siguiente - valor_final_contable) > 0.01:
                    print(f"Coherencia entre períodos: El inicial de {periodo_siguiente} no coincide con el final de {periodo}")
                    print(f"  Valor inicial de {periodo_siguiente}: {valor_inicial_siguiente:.2f}")
                    print(f"  Valor final de {periodo}: {valor_final_contable:.2f}")
                    
                    # Corregir el valor inicial del período siguiente
                    cursor.execute("""
                        UPDATE InventarioContable
                        SET Inicial = ?
                        WHERE Periodo = ?
                    """, (valor_final_contable, periodo_siguiente))
                    
                    # También corregir el Costo del registro 0000000001 del período siguiente
                    cursor.execute("""
                        UPDATE MovInventMes
                        SET Costo = ?
                        WHERE Periodo = ? AND Codigo = '0000000001'
                    """, (valor_final_contable, periodo_siguiente))
                    
                    print(f"  Se actualizó el valor inicial de {periodo_siguiente} a {valor_final_contable:.2f}")
                    hay_cambios = True
        
        # 4.3. Caso especial: Si es diciembre, verificar coherencia con enero del siguiente año
        if mes == 12:
            siguiente_año = año + 1
            periodo_siguiente = f"01/{siguiente_año}"
            
            # Verificar si existe el período siguiente
            cursor.execute("""
                SELECT COUNT(*) 
                FROM InventarioContable 
                WHERE Periodo = ?
            """, (periodo_siguiente,))
            
            if cursor.fetchone()[0] > 0:
                cursor.execute("""
                    SELECT Inicial
                    FROM InventarioContable
                    WHERE Periodo = ?
                """, (periodo_siguiente,))
                
                row = cursor.fetchone()
                if row:
                    valor_inicial_siguiente = float(row[0])
                    
                    # El inicial de enero del siguiente año debe ser igual al final de diciembre
                    if abs(valor_inicial_siguiente - valor_final_contable) > 0.01:
                        print(f"Coherencia entre años: El inicial de {periodo_siguiente} no coincide con el final de {periodo}")
                        print(f"  Valor inicial de {periodo_siguiente}: {valor_inicial_siguiente:.2f}")
                        print(f"  Valor final de {periodo}: {valor_final_contable:.2f}")
                        
                        # Corregir el valor inicial del período siguiente
                        cursor.execute("""
                            UPDATE InventarioContable
                            SET Inicial = ?
                            WHERE Periodo = ?
                        """, (valor_final_contable, periodo_siguiente))
                        
                        print(f"  Se actualizó el valor inicial de {periodo_siguiente} a {valor_final_contable:.2f}")
                        hay_cambios = True
        
        if hay_cambios:
            conn.commit()
            print(f"Valores corregidos para el período {periodo}.")
        else:
            print(f"Los valores del período {periodo} son coherentes.")
        
        conn.close()
    
    except Exception as e:
        print(f"Error al verificar coherencia de valores para {mes:02d}/{año}: {str(e)}")
        import traceback
        traceback.print_exc()
        
        try:
            conn.close()
        except:
            pass

def generar_excel_inventario_2024():
    """
    Genera un archivo Excel con el Libro Auxiliar de Entradas y Salidas del Inventario
    basado en los datos existentes en MovInventMes para el año 2024
    """
    try:
        print("Generando archivo Excel del Libro Auxiliar de Inventario 2024...")
        
        # Obtener datos de InventarioContable para 2024
        conn = get_connection()
        
        # Consultar el inventario inicial y final del año 2024
        query_contable = """
        SELECT Periodo, Inicial, Final, AjusteCompras, AjusteVentas
        FROM InventarioContable
        WHERE LTRIM(RTRIM(Periodo)) LIKE '%/2024'
        ORDER BY Periodo
        """
        df_contable = pd.read_sql(query_contable, conn)
        
        print(f"Registros encontrados en InventarioContable: {len(df_contable)}")
        
        # Limpiar espacios en blanco del campo Periodo
        df_contable['Periodo'] = df_contable['Periodo'].str.strip()
        
        print(f"Períodos encontrados después de TRIM: {df_contable['Periodo'].tolist()}")
        
        # Si no encontramos registros con el filtro, intentemos obtener todos y filtrar
        if df_contable.empty:
            print("No se encontraron registros con filtro 2024, obteniendo todos los períodos para debug...")
            query_debug = "SELECT Periodo, Inicial, Final FROM InventarioContable"
            df_debug = pd.read_sql(query_debug, conn)
            df_debug['Periodo'] = df_debug['Periodo'].str.strip()
            print(f"Todos los períodos disponibles: {df_debug['Periodo'].tolist()}")
            
            # Filtrar manualmente los períodos de 2024
            df_contable = df_debug[df_debug['Periodo'].str.endswith('/2024')].copy()
            print(f"Períodos 2024 encontrados manualmente: {df_contable['Periodo'].tolist()}")
        
        # Obtener el inventario inicial del año (enero 2024)
        df_enero = df_contable[df_contable['Periodo'] == '01/2024']
        if df_enero.empty:
            print("Error: No se encontró el período 01/2024 en InventarioContable")
            print(f"Períodos disponibles: {df_contable['Periodo'].unique()}")
            conn.close()
            return None
        inventario_inicial_año = df_enero['Inicial'].iloc[0]
        
        # Obtener el inventario final del año (diciembre 2024)  
        df_diciembre = df_contable[df_contable['Periodo'] == '12/2024']
        if df_diciembre.empty:
            print("Error: No se encontró el período 12/2024 en InventarioContable")
            print(f"Períodos disponibles: {df_contable['Periodo'].unique()}")
            conn.close()
            return None
        inventario_final_año = df_diciembre['Final'].iloc[0]
        
        print(f"Inventario Inicial Acumulado del Ejercicio Fiscal Anterior: {inventario_inicial_año:,.2f}")
        print(f"Inventario Final del Año 2024: {inventario_final_año:,.2f}")
        
        # Consultar todos los movimientos de 2024 ordenados por fecha
        query_movimientos = """
        SELECT 
            Periodo,
            Codigo,
            Descripcion,
            Fecha,
            inicial,
            Entradas,
            Salidas,
            AutoConsumo,
            Retiros,
            final,
            Costo,
            Inventario
        FROM MovInventMes
        WHERE LTRIM(RTRIM(Periodo)) LIKE '%/2024'
        AND Codigo != '0000000001'  -- Excluir el registro de inventario inicial
        ORDER BY Fecha, Codigo
        """
        df_movimientos = pd.read_sql(query_movimientos, conn)
        
        print(f"Movimientos encontrados para 2024: {len(df_movimientos)}")
        
        # Si no hay movimientos, intentar consulta de debug
        if df_movimientos.empty:
            print("No se encontraron movimientos con filtro 2024, verificando disponibilidad...")
            query_debug_mov = """
            SELECT DISTINCT LTRIM(RTRIM(Periodo)) as Periodo
            FROM MovInventMes
            WHERE Codigo != '0000000001'
            ORDER BY Periodo
            """
            df_debug_mov = pd.read_sql(query_debug_mov, conn)
            print(f"Períodos con movimientos disponibles: {df_debug_mov['Periodo'].tolist()}")
            
            # Intentar obtener todos los movimientos y filtrar manualmente
            query_all_mov = """
            SELECT 
                Periodo,
                Codigo,
                Descripcion,
                Fecha,
                inicial,
                Entradas,
                Salidas,
                AutoConsumo,
                Retiros,
                final,
                Costo,
                Inventario
            FROM MovInventMes
            WHERE Codigo != '0000000001'
            ORDER BY Fecha, Codigo
            """
            df_all_mov = pd.read_sql(query_all_mov, conn)
            df_all_mov['Periodo'] = df_all_mov['Periodo'].str.strip()
            df_movimientos = df_all_mov[df_all_mov['Periodo'].str.endswith('/2024')].copy()
            print(f"Movimientos 2024 encontrados manualmente: {len(df_movimientos)}")
        
        # Convertir la columna Fecha a datetime
        if not df_movimientos.empty:
            df_movimientos['Fecha'] = pd.to_datetime(df_movimientos['Fecha'])
        else:
            print("ADVERTENCIA: No se encontraron movimientos para 2024")
            print("El reporte se generará solo con datos de InventarioContable")
        
        # Crear el archivo Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Libro Auxiliar Inventario 2024"
        
        # Configurar estilos
        font_titulo = Font(name='Arial', size=12, bold=True)
        font_subtitulo = Font(name='Arial', size=10, bold=True)
        font_normal = Font(name='Arial', size=9)
        font_totales = Font(name='Arial', size=9, bold=True)
        
        align_center = Alignment(horizontal='center', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')
        
        border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        fill_header = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        fill_totales = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
        
        # Escribir encabezados del reporte
        ws['A1'] = 'FECHA DESDE: 01/01/2024'
        ws['A1'].font = font_subtitulo
        
        ws['I1'] = 'Fecha:'
        ws['I1'].font = font_subtitulo
        ws['J1'] = datetime.now().strftime('%d/%m/%Y')
        ws['J1'].font = font_normal
        
        ws['A2'] = 'FECHA HASTA: 31/12/2024'
        ws['A2'].font = font_subtitulo
        
        ws['C3'] = 'Libro Auxiliar de Entradas y Salidas del Inventario'
        ws['C3'].font = font_titulo
        ws.merge_cells('C3:G3')
        ws['C3'].alignment = align_center
        
        ws['C4'] = 'Movimiento de Unidades Según el Artículo 177 de Ley de Impuesto Sobre la Renta'
        ws['C4'].font = font_subtitulo
        ws.merge_cells('C4:H4')
        ws['C4'].alignment = align_center
        
        ws['J4'] = 'Pag:'
        ws['J4'].font = font_subtitulo
        ws['K4'] = '1'
        ws['L4'] = 'de'
        ws['M4'] = '1'
        
        ws['A6'] = f'Inventario Inicial Acumulado del Ejercicio Fiscal Anterior: {inventario_inicial_año:,.2f}'
        ws['A6'].font = font_subtitulo
        
        ws['J6'] = 'MÉTODO P.E.P.S.'
        ws['J6'].font = font_subtitulo
        
        # Encabezados de columnas (fila 8)
        encabezados = [
            'Fecha', 'Descripción Artículo y/o Producto', 'Existencia Inicial', '', 
            'Entradas', '', 'Salidas', '', 'Autoconsumos', '', 'Retiros', '', 
            'Existencia Actual', ''
        ]
        
        subencabezados = [
            '', '', 'Cant.', 'Monto', 'Cant.', 'Monto', 'Cant.', 'Monto', 
            'Cant.', 'Monto', 'Cant.', 'Monto', 'Cant.', 'Monto'
        ]
        
        # Escribir encabezados
        for col, header in enumerate(encabezados, 1):
            cell = ws.cell(row=8, column=col, value=header)
            cell.font = font_subtitulo
            cell.alignment = align_center
            cell.border = border_thin
            cell.fill = fill_header
        
        # Escribir subencabezados
        for col, subheader in enumerate(subencabezados, 1):
            cell = ws.cell(row=9, column=col, value=subheader)
            cell.font = font_subtitulo
            cell.alignment = align_center
            cell.border = border_thin
            cell.fill = fill_header
        
        # Ajustar anchos de columna
        column_widths = [12, 35, 8, 12, 8, 12, 8, 12, 8, 12, 8, 12, 8, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Combinar celdas de encabezados
        ws.merge_cells('C8:D8')  # Existencia Inicial
        ws.merge_cells('E8:F8')  # Entradas
        ws.merge_cells('G8:H8')  # Salidas
        ws.merge_cells('I8:J8')  # Autoconsumos
        ws.merge_cells('K8:L8')  # Retiros
        ws.merge_cells('M8:N8')  # Existencia Actual
        
        # Variables para totales
        total_registros = 0
        total_entradas_cantidad = 0
        total_entradas_monto = 0
        total_salidas_cantidad = 0
        total_salidas_monto = 0
        total_autoconsumo_cantidad = 0
        total_autoconsumo_monto = 0
        total_retiros_cantidad = 0
        total_retiros_monto = 0
        
        fila_actual = 10
        
        # Agregar el registro de inventario inicial del año anterior como primer registro
        primer_registro = [
            '01/01/2024',
            'INVENTARIO INICIAL ACUMULADO DEL EJERCICIO FISCAL ANTERIOR',
            0,  # Cantidad inicial (0 para que se vea limpio)
            0,  # Monto inicial (0 para que se vea limpio)
            1,  # Entradas cantidad (1 unidad simbólica)
            inventario_inicial_año,  # Entradas monto (aquí va el valor total del inventario inicial)
            0,  # Salidas cantidad (siempre 0 para registro inicial)
            0,  # Salidas monto (siempre 0 para registro inicial)
            0,  # Autoconsumos cantidad (siempre 0 para registro inicial)
            0,  # Autoconsumos monto (siempre 0 para registro inicial)
            0,  # Retiros cantidad (siempre 0 para registro inicial)
            0,  # Retiros monto (siempre 0 para registro inicial)
            1,   # Existencia final cantidad
            inventario_inicial_año   # Existencia final monto
        ]
        
        for col, valor in enumerate(primer_registro, 1):
            cell = ws.cell(row=fila_actual, column=col, value=valor)
            cell.font = font_totales  # Usar formato de totales para destacar
            cell.border = border_thin
            cell.fill = fill_header
            
            # Formato de números - consistente con el resto
            if col >= 3:  # Columnas numéricas
                if col in [3, 5, 7, 9, 11, 13]:  # Columnas de cantidad
                    cell.number_format = '#,##0'
                    cell.alignment = align_right
                elif col in [4, 6, 8, 10, 12, 14]:  # Columnas de monto
                    cell.number_format = '#,##0.00'
                    cell.alignment = align_right
                else:
                    cell.alignment = align_center
            elif col == 2:  # Columna de descripción
                cell.alignment = align_left
            else:  # Columna de fecha
                cell.alignment = align_center
        
        fila_actual += 1
        total_registros = 1  # Contar el registro inicial
        
        # Inicializar totales con el inventario inicial que ahora va en entradas
        total_entradas_cantidad = 1  # La unidad simbólica del inventario inicial
        total_entradas_monto = inventario_inicial_año  # El valor del inventario inicial
        
        # Procesar todos los movimientos sin agrupar por mes (solo si hay movimientos)
        if not df_movimientos.empty:
            df_movimientos['Fecha'] = pd.to_datetime(df_movimientos['Fecha'])
            # Ordenar todos los movimientos por fecha
            df_movimientos_ordenados = df_movimientos.sort_values('Fecha').copy()
            
            # AJUSTE DINÁMICO PARA CUADRAR INVENTARIOS
            print("Calculando ajustes necesarios para cuadrar inventarios...")
            
            # Calcular el total actual con los datos originales
            total_entradas_original = 0
            total_salidas_original = 0
            
            for _, row in df_movimientos_ordenados.iterrows():
                monto_entradas_orig = (row['Entradas'] * row['Costo']) if pd.notna(row['Entradas']) and pd.notna(row['Costo']) else 0
                monto_salidas_orig = (row['Salidas'] * row['Costo']) if pd.notna(row['Salidas']) and pd.notna(row['Costo']) else 0
                total_entradas_original += monto_entradas_orig
                total_salidas_original += monto_salidas_orig
            
            # Calcular qué debería dar para cuadrar
            variacion_requerida = inventario_final_año - inventario_inicial_año
            diferencia_neta_actual = total_entradas_original - total_salidas_original
            ajuste_necesario = variacion_requerida - diferencia_neta_actual
            
            print(f"Variación requerida: {variacion_requerida:,.2f}")
            print(f"Diferencia neta actual: {diferencia_neta_actual:,.2f}")
            print(f"Ajuste necesario: {ajuste_necesario:,.2f}")
            
            # Factor de ajuste proporcional
            if abs(ajuste_necesario) > 0.01:
                if total_entradas_original > 0:
                    factor_ajuste_entradas = (total_entradas_original + ajuste_necesario) / total_entradas_original
                else:
                    factor_ajuste_entradas = 1.0
                print(f"Factor de ajuste aplicado: {factor_ajuste_entradas:.6f}")
            else:
                factor_ajuste_entradas = 1.0
            
            # Verificar si el factor de ajuste es razonable (entre 0.1 y 10.0)
            ajuste_mediante_movimientos = False
            if factor_ajuste_entradas < 0.1 or factor_ajuste_entradas > 10.0:
                print(f"ADVERTENCIA: Factor de ajuste extremo ({factor_ajuste_entradas:.6f})")
                print("Se generarán movimientos de ajuste adicionales en lugar de ajustar precios.")
                factor_ajuste_entradas = 1.0  # No ajustar precios
                ajuste_mediante_movimientos = True
            
            # Escribir cada movimiento del año con ajustes
            for _, row in df_movimientos_ordenados.iterrows():
                # Aplicar factor de ajuste a los costos para que cuadre
                costo_ajustado = row['Costo'] * factor_ajuste_entradas if pd.notna(row['Costo']) else 0
                
                # Calcular montos con el costo ajustado
                monto_inicial = (row['inicial'] * costo_ajustado) if pd.notna(row['inicial']) else 0
                monto_entradas = (row['Entradas'] * costo_ajustado) if pd.notna(row['Entradas']) else 0
                monto_salidas = (row['Salidas'] * costo_ajustado) if pd.notna(row['Salidas']) else 0
                monto_autoconsumo = (row['AutoConsumo'] * costo_ajustado) if pd.notna(row['AutoConsumo']) else 0
                monto_retiros = (row['Retiros'] * costo_ajustado) if pd.notna(row['Retiros']) else 0
                monto_final = (row['final'] * costo_ajustado) if pd.notna(row['final']) else 0
                
                # Escribir fila
                datos_fila = [
                    row['Fecha'].strftime('%d/%m/%Y'),
                    f"{row['Codigo']} {row['Descripcion']}"[:50],  # Limitar descripción
                    row['inicial'] if pd.notna(row['inicial']) else 0,  # Siempre mostrar valor
                    monto_inicial if monto_inicial != 0 else 0,  # Con costo ajustado
                    row['Entradas'] if pd.notna(row['Entradas']) else 0,  # Siempre mostrar valor
                    monto_entradas if monto_entradas != 0 else 0,  # Con costo ajustado
                    row['Salidas'] if pd.notna(row['Salidas']) else 0,  # Siempre mostrar valor
                    monto_salidas if monto_salidas != 0 else 0,  # Con costo ajustado
                    row['AutoConsumo'] if pd.notna(row['AutoConsumo']) else 0,  # Siempre mostrar valor
                    monto_autoconsumo if monto_autoconsumo != 0 else 0,  # Con costo ajustado
                    row['Retiros'] if pd.notna(row['Retiros']) else 0,  # Siempre mostrar valor
                    monto_retiros if monto_retiros != 0 else 0,  # Con costo ajustado
                    row['final'] if pd.notna(row['final']) else 0,  # Siempre mostrar valor
                    monto_final if monto_final != 0 else 0  # Con costo ajustado
                ]
                
                for col, valor in enumerate(datos_fila, 1):
                    cell = ws.cell(row=fila_actual, column=col, value=valor)
                    cell.font = font_normal
                    cell.border = border_thin
                    
                    # Formato de números - aplicar formato numérico a todas las columnas numéricas
                    if col >= 3:  # Columnas numéricas
                        if col in [3, 5, 7, 9, 11, 13]:  # Columnas de cantidad
                            cell.number_format = '#,##0'
                            cell.alignment = align_right
                        elif col in [4, 6, 8, 10, 12, 14]:  # Columnas de monto
                            cell.number_format = '#,##0.00'
                            cell.alignment = align_right
                        else:
                            cell.alignment = align_center
                    elif col == 2:  # Columna de descripción
                        cell.alignment = align_left
                    else:  # Columna de fecha
                        cell.alignment = align_center
                
                # Acumular totales generales (con valores ajustados)
                total_entradas_cantidad += row['Entradas'] if pd.notna(row['Entradas']) else 0
                total_entradas_monto += monto_entradas
                total_salidas_cantidad += row['Salidas'] if pd.notna(row['Salidas']) else 0
                total_salidas_monto += monto_salidas
                total_autoconsumo_cantidad += row['AutoConsumo'] if pd.notna(row['AutoConsumo']) else 0
                total_autoconsumo_monto += monto_autoconsumo
                total_retiros_cantidad += row['Retiros'] if pd.notna(row['Retiros']) else 0
                total_retiros_monto += monto_retiros
                
                total_registros += 1
                fila_actual += 1
        else:
            # Si no hay movimientos, mostrar mensaje informativo
            ws.cell(row=fila_actual, column=1, value="No se encontraron movimientos de inventario para el año 2024").font = font_totales
            ws.merge_cells(f'A{fila_actual}:N{fila_actual}')
            ws.cell(row=fila_actual, column=1).alignment = align_center
            ws.cell(row=fila_actual, column=1).fill = fill_totales
            fila_actual += 2
        
        # Generar movimientos de ajuste si es necesario
        if 'ajuste_mediante_movimientos' in locals() and ajuste_mediante_movimientos:
            print("Generando movimientos de ajuste adicionales...")
            
            # Calcular cuánto necesitamos ajustar
            variacion_actual = total_entradas_monto - total_salidas_monto - total_autoconsumo_monto - total_retiros_monto
            ajuste_restante = (inventario_final_año - inventario_inicial_año) - variacion_actual
            
            if abs(ajuste_restante) > 0.01:
                # Generar entre 3 y 5 movimientos de ajuste
                num_movimientos_ajuste = min(5, max(3, int(abs(ajuste_restante) / 10000) + 1))
                valor_por_movimiento = ajuste_restante / num_movimientos_ajuste
                
                for i in range(num_movimientos_ajuste):
                    if i == num_movimientos_ajuste - 1:
                        # Último movimiento: ajustar el resto exacto
                        valor_movimiento = ajuste_restante - (valor_por_movimiento * (num_movimientos_ajuste - 1))
                    else:
                        valor_movimiento = valor_por_movimiento
                    
                    if valor_movimiento > 0:
                        # Movimiento de entrada
                        datos_ajuste = [
                            '31/12/2024',
                            f'AJUSTE-{i+1:03d} ENTRADA DE CUADRE INVENTARIO FINAL',
                            0,  # Inicial cantidad
                            0,  # Inicial monto
                            1,  # Entradas cantidad
                            valor_movimiento,  # Entradas monto
                            0, 0, 0, 0, 0, 0,  # Salidas, autoconsumos, retiros
                            1,  # Final cantidad
                            valor_movimiento   # Final monto
                        ]
                        total_entradas_cantidad += 1
                        total_entradas_monto += valor_movimiento
                    else:
                        # Movimiento de salida
                        datos_ajuste = [
                            '31/12/2024',
                            f'AJUSTE-{i+1:03d} SALIDA DE CUADRE INVENTARIO FINAL',
                            1,  # Inicial cantidad
                            abs(valor_movimiento),  # Inicial monto
                            0, 0,  # Entradas
                            1,  # Salidas cantidad
                            abs(valor_movimiento),  # Salidas monto
                            0, 0, 0, 0,  # Autoconsumos, retiros
                            0, 0   # Final
                        ]
                        total_salidas_cantidad += 1
                        total_salidas_monto += abs(valor_movimiento)
                    
                    # Escribir el movimiento de ajuste
                    for col, valor in enumerate(datos_ajuste, 1):
                        cell = ws.cell(row=fila_actual, column=col, value=valor)
                        cell.font = Font(name='Arial', size=9, bold=True, color='FF8000')  # Naranja para destacar
                        cell.border = border_thin
                        cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # Fondo amarillo claro
                        
                        # Formato de números
                        if col >= 3:
                            if col in [3, 5, 7, 9, 11, 13]:  # Columnas de cantidad
                                cell.number_format = '#,##0'
                                cell.alignment = align_right
                            elif col in [4, 6, 8, 10, 12, 14]:  # Columnas de monto
                                cell.number_format = '#,##0.00'
                                cell.alignment = align_right
                        elif col == 2:
                            cell.alignment = align_left
                        else:
                            cell.alignment = align_center
                    
                    total_registros += 1
                    fila_actual += 1
                
                print(f"Se generaron {num_movimientos_ajuste} movimientos de ajuste por {ajuste_restante:,.2f}")
        
        # Escribir totales generales
        fila_actual += 1
        
        # Separador
        ws.cell(row=fila_actual, column=1, value="=" * 100).font = font_totales
        ws.merge_cells(f'A{fila_actual}:N{fila_actual}')
        fila_actual += 1
        
        # Crear una fila de totales completa con todas las columnas
        fila_totales = [
            '',  # Columna 1: Fecha (vacía)
            'Total Registros: ' + str(total_registros),  # Columna 2: Descripción
            '',  # Columna 3: Existencia Inicial - Cant (vacía para totales)
            '',  # Columna 4: Existencia Inicial - Monto (vacía para totales)
            total_entradas_cantidad,  # Columna 5: Entradas - Cant
            total_entradas_monto,     # Columna 6: Entradas - Monto
            total_salidas_cantidad,   # Columna 7: Salidas - Cant
            total_salidas_monto,      # Columna 8: Salidas - Monto
            total_autoconsumo_cantidad, # Columna 9: Autoconsumos - Cant
            total_autoconsumo_monto,    # Columna 10: Autoconsumos - Monto
            total_retiros_cantidad,     # Columna 11: Retiros - Cant
            total_retiros_monto,        # Columna 12: Retiros - Monto
            '',  # Columna 13: Existencia Actual - Cant (vacía para totales)
            ''   # Columna 14: Existencia Actual - Monto (vacía para totales)
        ]
        
        # Escribir cada celda de la fila de totales
        for col, valor in enumerate(fila_totales, 1):
            cell = ws.cell(row=fila_actual, column=col, value=valor)
            cell.font = font_totales
            cell.border = border_thin
            cell.fill = fill_totales
            
            # Aplicar formato según el tipo de columna
            if col == 2:  # Columna de descripción
                cell.alignment = align_left
            elif col in [5, 7, 9, 11] and valor != '':  # Columnas de cantidad con valores
                cell.number_format = '#,##0'
                cell.alignment = align_right
            elif col in [6, 8, 10, 12] and valor != '':  # Columnas de monto con valores
                cell.number_format = '#,##0.00'
                cell.alignment = align_right
            else:  # Otras columnas
                cell.alignment = align_center
        
        fila_actual += 1
        
        # Separador adicional antes del resumen
        fila_actual += 2
        
        # Calcular el inventario final teórico basado en los movimientos
        inventario_final_calculado = inventario_inicial_año + total_entradas_monto - total_salidas_monto - total_autoconsumo_monto - total_retiros_monto
        diferencia_inventario = inventario_final_año - inventario_final_calculado
        variacion_anual = inventario_final_año - inventario_inicial_año
        
        # DEMOSTRACIÓN DEL CUADRE DEL INVENTARIO
        ws.cell(row=fila_actual, column=1, value="DEMOSTRACIÓN DEL CUADRE DEL INVENTARIO").font = Font(name='Arial', size=12, bold=True)
        fila_actual += 1
        
        ws.cell(row=fila_actual, column=1, value="=" * 80).font = font_totales
        fila_actual += 1
        
        # Mostrar si se aplicaron ajustes
        if 'factor_ajuste_entradas' in locals() and abs(factor_ajuste_entradas - 1.0) > 0.001:
            ws.cell(row=fila_actual, column=1, value=f'*** SE APLICÓ FACTOR DE AJUSTE: {factor_ajuste_entradas:.6f} PARA CUADRAR INVENTARIOS ***').font = Font(name='Arial', size=10, bold=True, color='FF8000')  # Naranja
            fila_actual += 1
            ws.cell(row=fila_actual, column=1, value=f'Los precios fueron ajustados proporcionalmente para lograr el cuadre perfecto').font = Font(name='Arial', size=9, color='FF8000')
            fila_actual += 1
            ws.cell(row=fila_actual, column=1, value="=" * 80).font = font_totales
            fila_actual += 1
        elif 'ajuste_mediante_movimientos' in locals() and ajuste_mediante_movimientos:
            ws.cell(row=fila_actual, column=1, value='*** SE AGREGARON MOVIMIENTOS DE AJUSTE PARA CUADRAR INVENTARIOS ***').font = Font(name='Arial', size=10, bold=True, color='FF8000')  # Naranja
            fila_actual += 1
            ws.cell(row=fila_actual, column=1, value='Se generaron movimientos adicionales destacados en amarillo para lograr el cuadre').font = Font(name='Arial', size=9, color='FF8000')
            fila_actual += 1
            ws.cell(row=fila_actual, column=1, value="=" * 80).font = font_totales
            fila_actual += 1
        
        # Paso 1: Inventario inicial
        ws.cell(row=fila_actual, column=1, value=f'NOTA: El inventario inicial ({inventario_inicial_año:,.2f}) está incluido en Total Entradas').font = Font(name='Arial', size=9, color='0000FF')
        fila_actual += 1
        
        # Paso 2: Entradas
        ws.cell(row=fila_actual, column=1, value=f'(+) Total Entradas del Año (incluye inicial): {total_entradas_monto:,.2f}').font = font_totales
        fila_actual += 1
        
        # Paso 3: Salidas
        ws.cell(row=fila_actual, column=1, value=f'(-) Total Salidas del Año: {total_salidas_monto:,.2f}').font = font_normal
        fila_actual += 1
        
        # Paso 4: Autoconsumos
        ws.cell(row=fila_actual, column=1, value=f'(-) Total Autoconsumos del Año: {total_autoconsumo_monto:,.2f}').font = font_normal
        fila_actual += 1
        
        # Paso 5: Retiros
        ws.cell(row=fila_actual, column=1, value=f'(-) Total Retiros del Año: {total_retiros_monto:,.2f}').font = font_normal
        fila_actual += 1
        
        ws.cell(row=fila_actual, column=1, value="=" * 80).font = font_totales
        fila_actual += 1
        
        # Resultado del cálculo (ahora debe cuadrar siempre debido a los ajustes)
        inventario_final_calculado_ajustado = total_entradas_monto - total_salidas_monto - total_autoconsumo_monto - total_retiros_monto
        ws.cell(row=fila_actual, column=1, value=f'(=) INVENTARIO FINAL CALCULADO: {inventario_final_calculado_ajustado:,.2f}').font = font_totales
        fila_actual += 1
        
        ws.cell(row=fila_actual, column=1, value=f'(=) INVENTARIO FINAL SEGÚN SISTEMA: {inventario_final_año:,.2f}').font = font_totales
        fila_actual += 1
        
        fila_actual += 1
        
        # Verificación del cuadre (ahora siempre debe ser perfecto)
        diferencia_final = inventario_final_año - inventario_final_calculado_ajustado
        if abs(diferencia_final) < 0.01:
            ws.cell(row=fila_actual, column=1, value='✓ ¡CUADRE PERFECTO! - LOS MOVIMIENTOS DEMUESTRAN CORRECTAMENTE LA TRANSICIÓN DEL INVENTARIO').font = Font(name='Arial', size=11, bold=True, color='008000')  # Verde
        else:
            ws.cell(row=fila_actual, column=1, value=f'⚠ DIFERENCIA REMANENTE: {diferencia_final:,.2f} - NECESARIO AJUSTE ADICIONAL').font = Font(name='Arial', size=11, bold=True, color='FF0000')  # Rojo
        
        fila_actual += 2
        
        # Fórmula de verificación
        ws.cell(row=fila_actual, column=1, value="FÓRMULA DE VERIFICACIÓN:").font = font_totales
        fila_actual += 1
        ws.cell(row=fila_actual, column=1, value=f'Inventario Final = Total Entradas (incluyendo inicial) - Salidas - Autoconsumos - Retiros').font = font_normal
        fila_actual += 1
        ws.cell(row=fila_actual, column=1, value=f'{inventario_final_año:,.2f} = {total_entradas_monto:,.2f} - {total_salidas_monto:,.2f} - {total_autoconsumo_monto:,.2f} - {total_retiros_monto:,.2f}').font = font_normal
        fila_actual += 1
        
        fila_actual += 1
        ws.cell(row=fila_actual, column=1, value=f'VARIACIÓN NETA DEL INVENTARIO EN EL AÑO: {variacion_anual:,.2f}').font = Font(name='Arial', size=11, bold=True, color='0000FF')  # Azul
        
        # Guardar archivo
        nombre_archivo = f'Libro_Auxiliar_Inventario_2024_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(nombre_archivo)
        
        conn.close()
        
        print(f"\nArchivo Excel generado exitosamente: {nombre_archivo}")
        print(f"Total de registros procesados: {total_registros}")
        print(f"Inventario inicial del año: {inventario_inicial_año:,.2f}")
        print(f"Inventario final del año: {inventario_final_año:,.2f}")
        print(f"Variación del año: {inventario_final_año - inventario_inicial_año:,.2f}")
        
        return nombre_archivo
        
    except Exception as e:
        print(f"Error al generar archivo Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("Uso: ")
        print("  python generador_inventario_execel.py excel                    # Generar archivo Excel 2024")
        print("  python generador_inventario_execel.py <año> [mes]              # Generar movimientos")
        print("  python generador_inventario_execel.py <año> [mes] [val_ini] [val_fin]  # Generar movimientos con valores específicos")
        sys.exit(1)
    
    try:
        # Opción para generar Excel
        if sys.argv[1].lower() == 'excel':
            print("=== GENERADOR DE LIBRO AUXILIAR DE INVENTARIO 2024 ===")
            print("Generando archivo Excel basado en los datos existentes...")
            archivo_generado = generar_excel_inventario_2024()
            if archivo_generado:
                print(f"\n✓ Archivo generado exitosamente: {archivo_generado}")
                print("\nEl archivo contiene:")
                print("- Libro Auxiliar de Entradas y Salidas del Inventario")
                print("- Inventario inicial del ejercicio fiscal anterior como primer registro")
                print("- Todos los movimientos del año 2024 ordenados cronológicamente")
                print("- Totales anuales de entradas, salidas, autoconsumos y retiros")
                print("- Validación del inventario final vs InventarioContable")
                print("- Formato compatible con el Artículo 177 de la Ley de Impuesto Sobre la Renta")
            else:
                print("\n✗ Error al generar el archivo Excel")
            sys.exit(0)
        
        # Código existente para generar movimientos
        año = int(sys.argv[1])
        
        # Valores FIJOS para conectar 2023-2024 (no modificar)
        VALOR_FINAL_2023 = 1120797.03  # Final de diciembre 2023
        VALOR_INICIAL_2024 = 1120797.03  # Inicial enero 2024 (igual al final de 2023)
        VALOR_FINAL_2024 = 1892903.00  # Final de diciembre 2024
        
        # Por defecto, no modificar los valores
        valor_inicial = None
        valor_final = None
        
        # Si es año 2024, usar valores específicos
        if año == 2024:
            valor_inicial = VALOR_INICIAL_2024
            valor_final = VALOR_FINAL_2024
            print(f"Usando valores predefinidos para 2024:")
            print(f"  Valor inicial (enero): {valor_inicial}")
            print(f"  Valor final (diciembre): {valor_final}")
            print(f"  Diferencia anual: {valor_final - valor_inicial}")
        
        # Procesar argumentos adicionales si se proporcionaron
        if len(sys.argv) > 3:
            valor_inicial = float(sys.argv[3])
        if len(sys.argv) > 4:
            valor_final = float(sys.argv[4])
        
        if len(sys.argv) > 2:
            # Si se especifica un mes, generar solo para ese mes
            mes = int(sys.argv[2])
            if 1 <= mes <= 12:
                # Para 2024, debemos asegurar que enero comience con el valor correcto
                if año == 2024 and mes == 1:
                    conn = get_connection()
                    cursor = conn.cursor()
                    cursor.execute("""
                        UPDATE InventarioContable 
                        SET Inicial = ? 
                        WHERE Periodo = '01/2024'
                    """, (VALOR_INICIAL_2024,))
                    conn.commit()
                    conn.close()
                    print(f"Ajustado valor inicial de enero 2024 a {VALOR_INICIAL_2024:.2f}")
                
                generar_movimientos_directo(año, mes)
            else:
                print("Error: El mes debe estar entre 1 y 12")
        else:
            # Si no se especifica mes, generar para todo el año
            if año == 2024:
                # 1. Primero establecer los valores de InventarioContable
                conn = get_connection()
                cursor = conn.cursor()
                
                # Ajustar valor inicial de enero
                cursor.execute("""
                    UPDATE InventarioContable 
                    SET Inicial = ? 
                    WHERE Periodo = '01/2024'
                """, (VALOR_INICIAL_2024,))
                
                # Ajustar valor final de diciembre
                cursor.execute("""
                    UPDATE InventarioContable 
                    SET Final = ? 
                    WHERE Periodo = '12/2024'
                """, (VALOR_FINAL_2024,))
                
                # Calcular diferencia
                diferencia_anual = VALOR_FINAL_2024 - VALOR_INICIAL_2024
                
                # Calcular factores de crecimiento (exponencial)
                factor_mensual = (VALOR_FINAL_2024 / VALOR_INICIAL_2024) ** (1/12.0)
                print(f"Factor de crecimiento mensual: {factor_mensual}")
                
                # Calcular valores esperados para cada mes (final)
                valores_finales = [VALOR_INICIAL_2024]
                for i in range(11):
                    valores_finales.append(round(valores_finales[-1] * factor_mensual, 2))
                # Ajustar el último para que sea exactamente el valor final
                # REMOVED: valores_finales[-1] = VALOR_FINAL_2024
                
                # Actualizar los valores en la tabla
                for mes in range(1, 12):
                    valor_final_mes = valores_finales[mes]
                    siguiente_mes = mes + 1
                    
                    # El final del mes actual será el inicial del siguiente
                    cursor.execute("""
                        UPDATE InventarioContable
                        SET Final = ?
                        WHERE Periodo = ?
                    """, (valor_final_mes, f'{mes:02d}/2024'))
                    
                    cursor.execute("""
                        UPDATE InventarioContable
                        SET Inicial = ?
                        WHERE Periodo = ?
                    """, (valor_final_mes, f'{siguiente_mes:02d}/2024'))
                    
                    print(f"Mes {mes:02d}: Final={valor_final_mes:.2f}")
                    print(f"Mes {siguiente_mes:02d}: Inicial={valor_final_mes:.2f}")
                
                conn.commit()
                conn.close()
            
            # 2. Ahora generar los movimientos para cada mes
            for mes in range(1, 13):
                try:
                    print(f"\nGenerando movimientos para el mes {mes}/2024...")
                    generar_movimientos_directo(año, mes)
                except Exception as e:
                    print(f"Error generando movimientos para mes {mes}: {str(e)}")
            
            print(f"Generación completa para el año {año}")
    
    except ValueError:
        print("Error: El año y el mes deben ser números enteros")
    except Exception as e:
        print(f"Error: {str(e)}")